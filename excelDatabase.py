import pyodbc
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import pandas as pd
import numpy as np
from excelFormatter import ExcelFormatter

class ExcelDatabase:
    def __init__(self, driver, server, database, UID="", PWD=""):
        self.driver = driver
        self.server = server
        self.database = database
        self.UID = UID
        self.PWD = PWD

    def connect(self):
        try:
            self.db = pyodbc.connect(f"""Driver={self.driver};
                                    Server={self.server}; Database={self.database};
                                    UID={self.UID}; PWD={self.PWD};
                                    Trusted_Connection=yes;""")
            print('Connection successful')
            return True
        except pyodbc.Error as e:
            print(e)
            return False
    
    def get_data(self, query):
        try:
            self.df = pd.read_sql_query(query, self.db)
            print('Data retrieved')
            return True
        except pd.errors.DatabaseError as e:
            print(e)
            return False
        
    def write_data(self, file_name, working_sheet_name):
        try:
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='w') as writer:
                self.df.to_excel(writer, sheet_name=working_sheet_name, index=False)
            print('Data written')
            return True
        except Exception as e:
            print(e)
            return False
        
    def generic_format(self, file_name, working_sheet_name, save_name=None):
        lastColumn = openpyxl.utils.get_column_letter(self.df.shape[1])
        self.get_column_length()
        try:
            wb = openpyxl.load_workbook(file_name)
            ws = wb[working_sheet_name]
            # Set column widths to fit
            for i in range(len(self.column_length)):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = self.column_length[i]
            # Generic Column Formatting
            ExcelFormatter.formatAxis(ws, (1, lastColumn), 1, Font(), PatternFill(), 
                                        Border(left=Side(border_style="thin", color="00000000"), right=Side(border_style="thin", color="00000000"), 
                                            top=Side(border_style="thin", color="00000000"), bottom=Side(border_style="thin", color="00000000")),
                                        Alignment(horizontal="left", vertical="bottom"))
            # Generic Header Formatting
            ExcelFormatter.formatCellRange(ws, f'A1:{lastColumn}1', Font(bold=True), PatternFill(),
                                Border(left=Side(border_style="thin", color="00000000"), right=Side(border_style="thin", color="00000000"), 
                                         top=Side(border_style="thin", color="00000000"), bottom=Side(border_style="thin", color="00000000")),
                                Alignment(horizontal="center", vertical="center"))
            if save_name is not None:
                wb.save(save_name)
            else:
                wb.save(file_name)
            wb.close()
            print('Formatting complete')
            return True
        except Exception as e:
            print(e)
            return False
        
    def get_column_length(self):
        measurer = np.vectorize(len)
        self.column_length = measurer(self.df.values.astype(str)).max(axis=0)
        for i in range(len(self.column_length)):
            self.column_length[i] = max(self.column_length[i], len(self.df.columns[i]))