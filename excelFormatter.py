import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment

class ExcelFormatter:
    def formatCellRange(ws, cellRange, ft=Font(), fl=PatternFill(), bd=Border(), al=Alignment()):
        # Formats all cells in a range with the same formatting
        for row in ws[cellRange]:
            for cell in row:
                cell.font = ft
                cell.fill = fl
                cell.border = bd
                cell.alignment = al
            
    def formatAxis(ws, cellRange, axis=0, ft=Font(), fl=PatternFill(), bd=Border(), al=Alignment()):
        # CellRange is a tuple (start, end); end is exclusive
        # Formats each row (default) or column in a range with the same formatting
        if not isinstance(cellRange, tuple):
            if axis == 0:
                ws.row_dimensions[cellRange].font = ft
                ws.row_dimensions[cellRange].fill = fl
                ws.row_dimensions[cellRange].border = bd
                ws.row_dimensions[cellRange].alignment = al
            else:
                if isinstance(cellRange, int):
                    cellRange = openpyxl.utils.get_column_letter(cellRange)
                ws.column_dimensions[cellRange].font = ft
                ws.column_dimensions[cellRange].fill = fl
                ws.column_dimensions[cellRange].border = bd
                ws.column_dimensions[cellRange].alignment = al
        else:
            if axis == 0:
                for row in range(cellRange[0], cellRange[1]):
                    ws.row_dimensions[row].font = ft
                    ws.row_dimensions[row].fill = fl
                    ws.row_dimensions[row].border = bd
                    ws.row_dimensions[row].alignment = al
            else:
                for column in openpyxl.utils.get_column_interval(cellRange[0], cellRange[1]):
                    ws.column_dimensions[column].font = ft
                    ws.column_dimensions[column].fill = fl
                    ws.column_dimensions[column].border = bd
                    ws.column_dimensions[column].alignment = al
                
    def setColumnWidth(ws, col, width):
        if not isinstance(col, tuple):
            if isinstance(col, int):
                col = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col].width = width
        else:
            if isinstance(width, int):
                for column in openpyxl.utils.get_column_interval(col[0], col[1]):
                    ws.column_dimensions[column].width = width
            else:
                for column, w in zip(openpyxl.utils.get_column_interval(col[0], col[1]), width):
                    ws.column_dimensions[column].width = w
            
    def fitColumnWidth(ws, cellRange):
        dim = {}
        for row in ws[cellRange]:
            for cell in row:
                dim[cell.column_letter] = max(dim.get(cell.column_letter, 0), len(str(cell.value)))
        for col, value in dim.items():
            ws.column_dimensions[col].width = value
            
    def formatCellType(ws, cellRange, format):
        for row in ws[cellRange]:
            for cell in row:
                cell.number_format = format
    
    def formatTextNumber(ws, cellRange, format):
        for row in ws[cellRange]:
            for cell in row:
                if isinstance(cell.value, str) and cell.value.isnumeric():
                    cell.value = int(cell.value)
                cell.number_format = format

    '''
# These are the constructors for the formatting objects
Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')

PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')

Border(left=Side(border_style=None, color='FF000000'),
        right=Side(border_style=None, color='FF000000'),
        top=Side(border_style=None, color='FF000000'),
        bottom=Side(border_style=None, color='FF000000'),
        diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
        outline=Side(border_style=None, color='FF000000'),
        vertical=Side(border_style=None, color='FF000000'),
        horizontal=Side(border_style=None, color='FF000000'))
        
Alignment(horizontal='general', vertical='bottom', text_rotation=0,
        wrap_text=False, shrink_to_fit=False, indent=0)
'''